from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from functools import wraps
import os
from datetime import datetime, timedelta
import csv
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from models import db, User, Department, Course, Classroom, Schedule, UnavailableTime, course_department
from sqlalchemy import inspect, text
import random

# =====================================================================================
# Ders Programı Yönetim Sistemi
# Bu sistem üniversite için bir ders programı yönetimi sağlar.
# Özellikler:
# - Bölüm ekleme, silme ve düzenleme
# - Ders ekleme, silme ve düzenleme
# - Derslik ekleme, silme ve düzenleme
# - Kullanıcı yönetimi (admin, öğretim görevlisi, öğrenci)
# - Ders programı oluşturma ve Excel'e aktarma
# =====================================================================================

# Göreceli yolları kullanarak dizinleri belirle
TEMPLATE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates')
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'ders_programi.db')

# Flask uygulamasını oluştur ve yapılandır
app = Flask(__name__, template_folder=TEMPLATE_DIR)
app.config['SECRET_KEY'] = 'gizli-anahtar-buraya'  # Güvenlik için session anahtarı
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + DB_PATH  # SQLite veritabanı bağlantısı
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # Performans için takip özelliğini kapat

# Veritabanı ve giriş yöneticisini başlat
db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'  # Giriş yapılmadığında yönlendirilecek sayfa

# Flask-Login için kullanıcı yükleme fonksiyonu
@login_manager.user_loader
def load_user(user_id):
    """
    Flask-Login için kullanıcı kimliğinden kullanıcı nesnesini yükler
    :param user_id: Kullanıcı kimlik numarası
    :return: Kullanıcı nesnesi veya None
    """
    return User.query.get(int(user_id))

# Admin yetkisi gerektiren sayfalar için dekoratör
def admin_required(f):
    """
    Bir rotaya sadece admin kullanıcıların erişebilmesini sağlayan dekoratör
    :param f: Dekore edilecek fonksiyon
    :return: Dekore edilmiş fonksiyon
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Kullanıcı giriş yapmamış veya admin değilse erişimi engelle
        if not current_user.is_authenticated or current_user.role != 'admin':
            flash('Bu sayfaya erişim yetkiniz yok!', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Ana sayfa - Ders programına yönlendirir
@app.route('/')
def index():
    """
    Ana sayfa, kullanıcıyı ders programı görüntüleme sayfasına yönlendirir
    """
    return redirect(url_for('view_schedule'))

# Giriş sayfası
@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Kullanıcı giriş sayfası
    GET: Giriş formunu göster
    POST: Kullanıcı giriş bilgilerini kontrol et
    """
    if request.method == 'POST':
        # Form verilerini al
        username = request.form.get('username')
        password = request.form.get('password')
        # Kullanıcıyı veritabanında ara
        user = User.query.filter_by(username=username).first()
        
        # Kullanıcı varsa ve şifre doğruysa giriş yap
        if user and user.check_password(password):
            login_user(user)
            # Kullanıcı rolüne göre yönlendirme yap
            if user.role == 'instructor':
                return redirect(url_for('my_schedule'))
            else:
                return redirect(url_for('view_schedule'))
        
        # Giriş başarısızsa hata mesajı göster
        flash('Geçersiz kullanıcı adı veya şifre!', 'error')
    return render_template('login.html')

# Çıkış sayfası
@app.route('/logout')
@login_required  # Sadece giriş yapmış kullanıcılar çıkış yapabilir
def logout():
    """
    Kullanıcının sistemden çıkış yapmasını sağlar
    """
    logout_user()
    return redirect(url_for('login'))

# Bölümler sayfası
@app.route('/departments', methods=['GET', 'POST'])
@admin_required  # Sadece adminler bölüm ekleyip silebilir
def departments():
    """
    Bölüm yönetim sayfası
    GET: Bölüm listesini göster
    POST: Yeni bölüm ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        code = request.form.get('code')
        name = request.form.get('name')
        
        # Aynı kodla başka bölüm var mı kontrol et
        if Department.query.filter_by(code=code).first():
            flash('Bu bölüm kodu zaten kullanımda!', 'error')
            return redirect(url_for('departments'))
        
        # Yeni bölüm oluştur ve kaydet
        department = Department(code=code, name=name)
        db.session.add(department)
        db.session.commit()
        
        flash('Bölüm başarıyla eklendi!', 'success')
        return redirect(url_for('departments'))
    
    # Tüm bölümleri getir ve görüntüle
    departments = Department.query.all()
    return render_template('departments.html', departments=departments)

# Dersler sayfası
@app.route('/courses', methods=['GET', 'POST'])
@admin_required  # Sadece adminler ders ekleyip silebilir
def courses():
    """
    Ders yönetim sayfası
    GET: Ders listesini göster
    POST: Yeni ders ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        code = request.form.get('code')
        name = request.form.get('name')
        department_ids = request.form.getlist('department_ids')  # Çoklu bölüm seçimi
        theory = request.form.get('theory', 0)
        practice = request.form.get('practice', 0)
        credits = request.form.get('credits', 0)
        instructor_id = request.form.get('instructor_id') if request.form.get('instructor_id') else None
        semester = request.form.get('semester', 1)
        is_mandatory = 'is_mandatory' in request.form
        
        # Aynı kodla başka ders var mı kontrol et
        if Course.query.filter_by(code=code).first():
            flash('Bu ders kodu zaten kullanımda!', 'error')
            return redirect(url_for('courses'))
        
        # Seçilen bölümleri kontrol et
        if not department_ids:
            flash('En az bir bölüm seçmelisiniz!', 'error')
            return redirect(url_for('courses'))
        
        # Yeni ders oluştur
        course = Course(
            code=code, 
            name=name,
            theory=theory,
            practice=practice,
            credits=credits,
            instructor_id=instructor_id,
            semester=semester,
            is_mandatory=is_mandatory
        )
        
        # Bölümleri ekle
        for dept_id in department_ids:
            department = Department.query.get(dept_id)
            if department:
                course.departments.append(department)
        
        db.session.add(course)
        db.session.commit()
        
        flash('Ders başarıyla eklendi!', 'success')
        return redirect(url_for('courses'))
    
    # Gerekli verileri getir ve görüntüle
    courses = Course.query.all()
    departments = Department.query.all()
    instructors = User.query.filter_by(role='instructor').all()
    return render_template('courses.html', courses=courses, departments=departments, instructors=instructors)

# Derslikler sayfası
@app.route('/classrooms', methods=['GET', 'POST'])
@admin_required  # Sadece adminler derslik ekleyip silebilir
def classrooms():
    """
    Derslik yönetim sayfası
    GET: Derslik listesini göster
    POST: Yeni derslik ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        code = request.form.get('code')
        capacity = request.form.get('capacity')
        
        # Aynı kodla başka derslik var mı kontrol et
        if Classroom.query.filter_by(code=code).first():
            flash('Bu derslik kodu zaten kullanımda!', 'error')
            return redirect(url_for('classrooms'))
        
        # Yeni derslik oluştur ve kaydet
        classroom = Classroom(code=code, capacity=capacity)
        db.session.add(classroom)
        db.session.commit()
        
        flash('Derslik başarıyla eklendi!', 'success')
        return redirect(url_for('classrooms'))
    
    # Tüm derslikleri getir ve görüntüle
    classrooms = Classroom.query.all()
    return render_template('classrooms.html', classrooms=classrooms)

# Kullanıcılar sayfası
@app.route('/users', methods=['GET', 'POST'])
@admin_required  # Sadece adminler kullanıcı ekleyip silebilir
def users():
    """
    Kullanıcı yönetim sayfası
    GET: Kullanıcı listesini göster
    POST: Yeni kullanıcı ekle
    """
    if request.method == 'POST':
        # Form verilerini al
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        name = request.form.get('name')
        department_id = request.form.get('department_id') if request.form.get('department_id') else None
        extra_info = request.form.get('extra_info')
        
        # Aynı kullanıcı adıyla başka kullanıcı var mı kontrol et
        if User.query.filter_by(username=username).first():
            flash('Bu kullanıcı adı zaten kullanımda!', 'error')
            return redirect(url_for('users'))
        
        # Yeni kullanıcı oluştur ve kaydet
        user = User(
            username=username, 
            role=role,
            name=name,
            department_id=department_id
        )
        user.set_password(password)
        
        db.session.add(user)
        db.session.commit()
        
        flash('Kullanıcı başarıyla eklendi!', 'success')
        return redirect(url_for('users'))
    
    # Gerekli verileri getir ve görüntüle
    users = User.query.all()
    departments = Department.query.all()
    return render_template('users.html', users=users, departments=departments)

# Kullanıcı silme endpoint'i
@app.route('/users/delete/<int:user_id>', methods=['POST'])
@admin_required
def delete_user(user_id):
    """
    Belirtilen ID'ye sahip kullanıcıyı siler
    :param user_id: Silinecek kullanıcının ID'si
    """
    try:
        # Kendini silmeye çalışıyor mu kontrolü
        if current_user.id == user_id:
            flash('Kendi hesabınızı silemezsiniz!', 'error')
            return redirect(url_for('users'))
        
        # Kullanıcıyı bul
        user = User.query.get_or_404(user_id)
        
        # Admin silinmeye çalışılıyor ve başka admin var mı kontrolü
        if user.role == 'admin':
            admin_count = User.query.filter_by(role='admin').count()
            if admin_count <= 1:
                flash('Son admin kullanıcıyı silemezsiniz!', 'error')
                return redirect(url_for('users'))
        
        # Kullanıcıyı sil
        db.session.delete(user)
        db.session.commit()
        flash('Kullanıcı başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
        flash('Kullanıcı silinirken bir hata oluştu!', 'error')
    
    return redirect(url_for('users'))

# Ders programı görüntüleme sayfası
@app.route('/view_schedule')
@login_required  # Sadece giriş yapmış kullanıcılar görebilir
def view_schedule():
    """
    Ders programını görüntüleme sayfası
    Tüm dersleri, derslikleri ve ders programını gösterir
    """
    # Haftanın günleri
    days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    
    # Veritabanından gerekli verileri çek
    schedule_items = Schedule.query.all()
    courses = Course.query.order_by(Course.code).all()  # Dersleri kod sırasına göre sırala
    classrooms = Classroom.query.order_by(Classroom.code).all()  # Derslikleri kod sırasına göre sırala
    
    # Bölümleri bul
    blm_dept = Department.query.filter_by(code='BLM').first()
    yzm_dept = Department.query.filter_by(code='YZM').first()
    
    # Bölümlere göre dersleri ayır (bir ders birden fazla bölüme ait olabilir)
    blm_courses = []
    yzm_courses = []
    
    for course in courses:
        # Her dersin bölümlerini kontrol et
        dept_ids = [d.id for d in course.departments]
        
        if blm_dept and blm_dept.id in dept_ids:
            blm_courses.append(course)
        
        if yzm_dept and yzm_dept.id in dept_ids:
            yzm_courses.append(course)
    
    # Bölümlere göre programları ayır
    blm_schedule_items = []
    yzm_schedule_items = []
    
    for item in schedule_items:
        course = Course.query.get(item.course_id)
        if course:
            # Dersin bölümlerini kontrol et
            dept_ids = [d.id for d in course.departments]
            
            if blm_dept and blm_dept.id in dept_ids:
                blm_schedule_items.append(item)
            
            if yzm_dept and yzm_dept.id in dept_ids:
                yzm_schedule_items.append(item)
    
    # Debug için konsola bilgi yazdır
    print("\n=== Debug Bilgileri ===")
    print(f"Toplam ders sayısı: {len(courses)}")
    print(f"BLM ders sayısı: {len(blm_courses)}")
    print(f"YZM ders sayısı: {len(yzm_courses)}")
    print(f"Toplam derslik sayısı: {len(classrooms)}")
    print(f"BLM program öğesi sayısı: {len(blm_schedule_items)}")
    print(f"YZM program öğesi sayısı: {len(yzm_schedule_items)}")
    
    # Şablonu render et
    return render_template('view_schedule.html',
                         blm_schedule_items=blm_schedule_items,
                         yzm_schedule_items=yzm_schedule_items,
                         schedule_items=schedule_items,
                         courses=courses,
                         blm_courses=blm_courses,
                         yzm_courses=yzm_courses,
                         classrooms=classrooms,
                         days=days,
                         blm_dept=blm_dept,
                         yzm_dept=yzm_dept)

# Program ekle endpoint'i
@app.route('/schedule/add', methods=['GET', 'POST'])
@admin_required  # Sadece adminler program ekleyebilir
def add_schedule():
    """
    Ders programına yeni bir ders ekler
    GET: Yönlendirme yapar
    POST: Yeni programı kaydeder
    """
    try:
        if request.method == 'GET':
            return redirect(url_for('view_schedule'))
        
        # Form verilerini al
        course_id = request.form.get('course_id')
        classroom_id = request.form.get('classroom_id')
        day = request.form.get('day')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')

        # Debug için form verilerini yazdır
        print(f"\n=== Form Verileri ===")
        print(f"Ders ID: {course_id}")
        print(f"Derslik ID: {classroom_id}")
        print(f"Gün: {day}")
        print(f"Başlangıç: {start_time}")
        print(f"Bitiş: {end_time}")
        print("==================\n")

        # Seçilen dersin öğretim üyesini bul
        course = Course.query.get(course_id)
        if course and course.instructor_id:
            instructor = User.query.get(course.instructor_id)
            print(f"Ders öğretim üyesi: {instructor.name if instructor else 'Atanmamış'}")
            
            # Öğretim üyesinin bu gün ve saatte müsait olmama durumu var mı kontrol et
            unavailable_times = UnavailableTime.query.filter(
                UnavailableTime.instructor_id == course.instructor_id,
                UnavailableTime.day == day,
                ((UnavailableTime.start_time <= start_time) & (UnavailableTime.end_time > start_time)) | 
                ((UnavailableTime.start_time < end_time) & (UnavailableTime.end_time >= end_time)) |
                ((UnavailableTime.start_time >= start_time) & (UnavailableTime.end_time <= end_time))
            ).all()
            
            if unavailable_times:
                flash(f'Öğretim üyesi ({instructor.name}) bu zaman diliminde müsait değil!', 'error')
                return redirect(url_for('view_schedule'))
            
            # Öğretim üyesinin bu zaman diliminde başka dersi var mı kontrol et
            instructor_conflicts = Schedule.query.join(Course).filter(
                Schedule.day == day,
                Schedule.start_time < end_time,
                Schedule.end_time > start_time,
                Course.instructor_id == course.instructor_id
            ).all()
            
            if instructor_conflicts:
                conflict_details = []
                for conflict in instructor_conflicts:
                    conflict_course = Course.query.get(conflict.course_id)
                    conflict_classroom = Classroom.query.get(conflict.classroom_id)
                    conflict_details.append(f"{conflict_course.code} ({conflict_classroom.code}, {conflict.start_time}-{conflict.end_time})")
                
                # Öğretim üyesi çakışması varsa uyar
                conflict_message = ", ".join(conflict_details)
                print(f'Öğretim üyesi ({instructor.name}) başka derste meşgul', 'error')
                return redirect(url_for('view_schedule'))

        # Seçilen derslik ve zamanda başka ders var mı kontrol et
        classroom_conflicts = Schedule.query.filter(
            Schedule.day == day,
            Schedule.start_time < end_time,
            Schedule.end_time > start_time,
            Schedule.classroom_id == classroom_id
        ).all()
        
        if classroom_conflicts:
            # Derslik çakışması varsa uyar
            conflict_details = []
            for conflict in classroom_conflicts:
                conflict_course = Course.query.get(conflict.course_id)
                conflict_details.append(f"{conflict_course.code} ({conflict.start_time}-{conflict.end_time})")
            
            conflict_message = ", ".join(conflict_details)
            conflict_classroom = Classroom.query.get(classroom_id)
            flash(f'Derslik {conflict_classroom.code} bu saatte dolu: {conflict_message}', 'error')
            return redirect(url_for('view_schedule'))
        
        # Yeni program öğesi oluştur ve kaydet
        schedule_item = Schedule(
            course_id=course_id,
            classroom_id=classroom_id,
            day=day,
            start_time=start_time,
            end_time=end_time
        )
        
        db.session.add(schedule_item)
        db.session.commit()
        
        flash('Ders programı başarıyla güncellendi!', 'success')
        
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
        flash('Ders programı eklenirken bir hata oluştu!', 'error')
        
    return redirect(url_for('view_schedule'))

# Program sil endpoint'i
@app.route('/schedule/delete/<int:schedule_id>', methods=['POST'])
@admin_required  # Sadece adminler program silebilir
def delete_schedule(schedule_id):
    """
    Belirtilen ID'ye sahip program öğesini siler
    :param schedule_id: Silinecek program öğesinin ID'si
    """
    try:
        # Program öğesini bul ve sil
        schedule_item = Schedule.query.get_or_404(schedule_id)
        db.session.delete(schedule_item)
        db.session.commit()
        flash('Program öğesi başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Program öğesi silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('view_schedule'))

# Bölüm silme endpoint'i
@app.route('/departments/delete/<int:department_id>', methods=['POST'])
@admin_required  # Sadece adminler bölüm silebilir
def delete_department(department_id):
    """
    Belirtilen ID'ye sahip bölümü siler
    :param department_id: Silinecek bölümün ID'si
    """
    try:
        # Bölümün kullanıldığı dersleri kontrol et
        courses_in_department = Course.query.filter_by(department_id=department_id).count()
        users_in_department = User.query.filter_by(department_id=department_id).count()
        
        # İlişkili kayıtlar varsa silme
        if courses_in_department > 0 or users_in_department > 0:
            flash(f'Bu bölüm silinemez: {courses_in_department} ders ve {users_in_department} kullanıcı bu bölüme bağlı!', 'error')
            return redirect(url_for('departments'))
            
        # Bölümü bul ve sil
        department = Department.query.get_or_404(department_id)
        db.session.delete(department)
        db.session.commit()
        flash('Bölüm başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Bölüm silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('departments'))

# Ders silme endpoint'i
@app.route('/courses/delete/<int:course_id>', methods=['POST'])
@admin_required  # Sadece adminler ders silebilir
def delete_course(course_id):
    """
    Belirtilen ID'ye sahip dersi siler
    :param course_id: Silinecek dersin ID'si
    """
    try:
        # Dersin kullanıldığı program öğeleri var mı kontrol et
        schedule_count = Schedule.query.filter_by(course_id=course_id).count()
        
        # İlişkili kayıtlar varsa silme
        if schedule_count > 0:
            flash(f'Bu ders silinemez: {schedule_count} program öğesi bu derse bağlı!', 'error')
            return redirect(url_for('courses'))
            
        # Dersi bul ve sil
        course = Course.query.get_or_404(course_id)
        db.session.delete(course)
        db.session.commit()
        flash('Ders başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Ders silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('courses'))

# Ders düzenleme endpoint'i
@app.route('/courses/edit/<int:course_id>', methods=['GET', 'POST'])
@admin_required  # Sadece adminler ders düzenleyebilir
def edit_course(course_id):
    """
    Belirtilen ID'ye sahip dersi düzenler
    :param course_id: Düzenlenecek dersin ID'si
    GET: Düzenleme formunu göster
    POST: Değişiklikleri kaydet
    """
    # Düzenlenecek dersi getir
    course = Course.query.get_or_404(course_id)
    
    if request.method == 'POST':
        try:
            # Form verilerini al
            name = request.form.get('name')
            department_ids = request.form.getlist('department_ids')  # Çoklu bölüm seçimi
            theory = request.form.get('theory', 0)
            practice = request.form.get('practice', 0)
            credits = request.form.get('credits', 0)
            instructor_id = request.form.get('instructor_id') if request.form.get('instructor_id') else None
            semester = request.form.get('semester', 1)
            is_mandatory = 'is_mandatory' in request.form
            
            # Seçilen bölümleri kontrol et
            if not department_ids:
                flash('En az bir bölüm seçmelisiniz!', 'error')
                return redirect(url_for('edit_course', course_id=course_id))
            
            # Dersi güncelle
            course.name = name
            course.theory = theory
            course.practice = practice
            course.credits = credits
            course.instructor_id = instructor_id
            course.semester = semester
            course.is_mandatory = is_mandatory
            
            # Bölümleri güncelle - önce tüm bölümleri temizle, sonra yeniden ekle
            course.departments.clear()
            for dept_id in department_ids:
                department = Department.query.get(dept_id)
                if department:
                    course.departments.append(department)
            
            db.session.commit()
            flash('Ders başarıyla güncellendi!', 'success')
            return redirect(url_for('courses'))
        except Exception as e:
            # Hata durumunda logla ve kullanıcıya bildir
            flash('Ders güncellenirken bir hata oluştu!', 'error')
            print(f"\n=== Hata ===")
            print(f"Hata mesajı: {str(e)}")
            print("============\n")
    
    # Formda kullanılacak verileri getir
    departments = Department.query.all()
    instructors = User.query.filter_by(role='instructor').all()
    
    # Dersin şu anda seçili bölümlerini al
    course_department_ids = [dept.id for dept in course.departments]
    
    return render_template('edit_course.html', course=course, departments=departments, 
                          instructors=instructors, course_department_ids=course_department_ids)

# Derslik silme endpoint'i
@app.route('/classrooms/delete/<int:classroom_id>', methods=['POST'])
@admin_required  # Sadece adminler derslik silebilir
def delete_classroom(classroom_id):
    """
    Belirtilen ID'ye sahip dersliği siler
    :param classroom_id: Silinecek dersliğin ID'si
    """
    try:
        # Dersliğin kullanıldığı program öğeleri var mı kontrol et
        schedule_count = Schedule.query.filter_by(classroom_id=classroom_id).count()
        
        # İlişkili kayıtlar varsa silme
        if schedule_count > 0:
            flash(f'Bu derslik silinemez: {schedule_count} program öğesi bu dersliğe bağlı!', 'error')
            return redirect(url_for('classrooms'))
            
        # Dersliği bul ve sil
        classroom = Classroom.query.get_or_404(classroom_id)
        db.session.delete(classroom)
        db.session.commit()
        flash('Derslik başarıyla silindi!', 'success')
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        flash('Derslik silinirken bir hata oluştu!', 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print("============\n")
    
    return redirect(url_for('classrooms'))

# Derslik düzenleme endpoint'i
@app.route('/classrooms/edit/<int:classroom_id>', methods=['GET', 'POST'])
@admin_required  # Sadece adminler derslik düzenleyebilir
def edit_classroom(classroom_id):
    """
    Belirtilen ID'ye sahip dersliği düzenler
    :param classroom_id: Düzenlenecek dersliğin ID'si
    GET: Düzenleme formunu göster
    POST: Değişiklikleri kaydet
    """
    # Düzenlenecek dersliği getir
    classroom = Classroom.query.get_or_404(classroom_id)
    
    if request.method == 'POST':
        try:
            # Form verilerini al
            capacity = request.form.get('capacity')
            
            # Dersliği güncelle
            classroom.capacity = capacity
            
            db.session.commit()
            flash('Derslik başarıyla güncellendi!', 'success')
            return redirect(url_for('classrooms'))
        except Exception as e:
            # Hata durumunda logla ve kullanıcıya bildir
            flash('Derslik güncellenirken bir hata oluştu!', 'error')
            print(f"\n=== Hata ===")
            print(f"Hata mesajı: {str(e)}")
            print("============\n")
    
    return render_template('edit_classroom.html', classroom=classroom)

# Ders programını Excel'e aktarma endpoint'i
@app.route('/export_schedule', methods=['GET'])
@admin_required  # Sadece adminler programı dışa aktarabilir
def export_schedule():
    """
    Mevcut ders programını Excel formatında dışa aktarır
    """
    try:
        # Excel çalışma kitabı oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "Ders Programı"
        
        # Sütun genişliklerini ayarla
        ws.column_dimensions['A'].width = 15  # Günler için
        for col in range(2, 6):  # 1-4 sınıflar için
            ws.column_dimensions[chr(64 + col)].width = 30
            
        # Haftanın günleri
        days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
        
        # Stil tanımları
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        day_font = Font(bold=True)
        day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        day_alignment = Alignment(horizontal='center', vertical='center')
        
        cell_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # İnce kenarlık stili
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Başlık satırını hazırla - Sınıf seviyelerini ekle
        ws.cell(row=1, column=1, value="Gün/Sınıf").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=1).alignment = header_alignment
        ws.cell(row=1, column=1).border = thin_border
        
        # Sınıf seviyelerini başlıklara ekle (1. Sınıf, 2. Sınıf, vb.)
        for grade in range(1, 5):  # 1-4. sınıflar
            cell = ws.cell(row=1, column=grade+1, value=f"{grade}. Sınıf")
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # Gün satırlarını ekle
        for row, day in enumerate(days, start=2):
            # Gün adını ekle
            cell = ws.cell(row=row, column=1, value=day)
            cell.font = day_font
            cell.fill = day_fill
            cell.alignment = day_alignment
            cell.border = thin_border
            
            # Her sınıf seviyesi için bu günde olan programları bul
            for grade in range(1, 5):  # 1-4. sınıflar
                cell = ws.cell(row=row, column=grade+1, value="")
                cell.border = thin_border
                cell.alignment = cell_alignment
                
                # Bu sınıfın yarıyıllarını belirle (her sınıf 2 yarıyıl içerir)
                first_semester = (grade - 1) * 2 + 1
                second_semester = first_semester + 1
                semesters = [first_semester, second_semester]
                
                # Bu gün ve sınıf seviyesinde olan dersleri bul
                # BLM ve YZM bölümlerini birlikte göster
                blm_dept = Department.query.filter_by(code='BLM').first()
                yzm_dept = Department.query.filter_by(code='YZM').first()
                
                if blm_dept and yzm_dept:
                    # Bu sınıfın yarıyıllarındaki dersleri bul
                    blm_courses = Course.query.filter(
                        Course.semester.in_(semesters),
                        Course.departments.any(id=blm_dept.id)
                    ).all()
                    
                    yzm_courses = Course.query.filter(
                        Course.semester.in_(semesters),
                        Course.departments.any(id=yzm_dept.id)
                    ).all()
                    
                    # Tüm kurs ID'lerini birleştir
                    course_ids = [course.id for course in blm_courses + yzm_courses]
                    
                    if course_ids:
                        # Bu günde ve bu kurslarda olan programları bul
                        schedule_items = Schedule.query.filter(
                            Schedule.day == day,
                            Schedule.course_id.in_(course_ids)
                        ).order_by(Schedule.start_time).all()
                        
                        # Program varsa hücreye ekle
                        if schedule_items:
                            cell_text = []
                            for item in schedule_items:
                                course = Course.query.get(item.course_id)
                                if not course:
                                    continue
                                    
                                classroom = Classroom.query.get(item.classroom_id)
                                instructor = User.query.get(course.instructor_id) if course.instructor_id else None
                                
                                # Dersin bölümünü bul
                                dept = course.departments[0] if course.departments else None
                                dept_code = dept.code if dept else ''
                                
                                course_info = (
                                    f"{course.code} - {course.name} ({dept_code}, {course.semester}. Yarıyıl)\n"
                                    f"Derslik: {classroom.code if classroom else 'Belirtilmemiş'}\n"
                                    f"Saat: {item.start_time}-{item.end_time}"
                                )
                                
                                if instructor:
                                    course_info += f"\nÖğr. Üyesi: {instructor.name}"
                                    
                                cell_text.append(course_info)
                            
                            if cell_text:
                                cell.value = "\n\n".join(cell_text)
            
            # Satır yüksekliğini ayarla
            ws.row_dimensions[row].height = 150
            
        # Geçici dosya oluştur ve Excel'i kaydet
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp_path = tmp.name
            
        # Excel dosyasını kullanıcıya gönder
        return send_file(
            tmp_path,
            as_attachment=True,
            download_name='ders_programi.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        # Hata durumunda logla ve kullanıcıya bildir
        error_msg = f"Ders programı dışa aktarılırken bir hata oluştu: {str(e)}"
        flash(error_msg, 'error')
        print(f"\n=== Hata ===")
        print(f"Hata mesajı: {str(e)}")
        print(f"Hata türü: {type(e).__name__}")
        print("============\n")
        return redirect(url_for('view_schedule'))

@app.route('/unavailable_times', methods=['GET', 'POST'])
@login_required
def manage_unavailable_times():
    if current_user.role != 'instructor':
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        day = request.form.get('day')
        start_time = request.form.get('start_time')
        end_time = request.form.get('end_time')
        reason = request.form.get('reason')
        
        unavailable_time = UnavailableTime(
            instructor_id=current_user.id,
            day=day,
            start_time=start_time,
            end_time=end_time,
            reason=reason
        )
        
        db.session.add(unavailable_time)
        db.session.commit()
        flash('Müsait olmayan zaman başarıyla eklendi.', 'success')
        return redirect(url_for('manage_unavailable_times'))
    
    unavailable_times = UnavailableTime.query.filter_by(instructor_id=current_user.id).all()
    return render_template('unavailable_times.html', unavailable_times=unavailable_times)

@app.route('/unavailable_times/delete/<int:id>', methods=['POST'])
@login_required
def delete_unavailable_time(id):
    if current_user.role != 'instructor':
        flash('Bu işlem için yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    unavailable_time = UnavailableTime.query.get_or_404(id)
    if unavailable_time.instructor_id != current_user.id:
        flash('Bu kaydı silme yetkiniz yok.', 'error')
        return redirect(url_for('manage_unavailable_times'))
    
    db.session.delete(unavailable_time)
    db.session.commit()
    flash('Müsait olmayan zaman başarıyla silindi.', 'success')
    return redirect(url_for('manage_unavailable_times'))

def generate_schedule(term=None):
    """
    Otomatik ders programı oluşturma fonksiyonu
    term: "guz" veya "bahar" olabilir. Güz ise 1,3,5,7. yarıyıllar, Bahar ise 2,4,6,8. yarıyıllar.
    """
    try:
        # Debug modunu kapalı tut - çok fazla loglama olmasın
        debug_mode = False
        
        # Mevcut programı temizle
        Schedule.query.delete()
        
        # Haftanın günleri ve saatler
        days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
        
        # 3 saatlik ders blokları
        time_slots = [
            ('09:00', '11:50'),
            ('13:00', '15:50')
        ]
        
        # Güz veya Bahar dönemine göre işlenecek yarıyıllar
        if term == "guz":
            semesters = [1, 3, 5, 7]  # Güz dönemi yarıyılları
            term_name = "Güz"
        elif term == "bahar":
            semesters = [2, 4, 6, 8]  # Bahar dönemi yarıyılları
            term_name = "Bahar"
        else:
            # Dönem belirtilmemişse tüm yarıyıllar
            semesters = list(range(1, 9))
            term_name = "Tüm"
        
        # Bölümleri bul
        blm_dept = Department.query.filter_by(code='BLM').first()
        yzm_dept = Department.query.filter_by(code='YZM').first()
        
        if not blm_dept or not yzm_dept:
            return False, "BLM veya YZM bölümü bulunamadı!"
            
        # Seçilen yarıyıllardaki tüm dersleri al
        all_courses = Course.query.filter(Course.semester.in_(semesters)).all()
        
        # BLM ve YZM bölümlerine ait dersleri belirle
        # Artık bir ders birden fazla bölüme ait olabilir
        blm_courses = []
        yzm_courses = []
        
        for course in all_courses:
            # Her dersin bölümlerini kontrol et
            departments = [d.id for d in course.departments]
            if blm_dept.id in departments:
                blm_courses.append(course)
            if yzm_dept.id in departments:
                yzm_courses.append(course)
        
        classrooms = Classroom.query.all()
        unavailable_times = UnavailableTime.query.all()
        
        print(f"\n=== {term_name} Dönemi Programı Oluşturuluyor ===")
        print(f"İşlenecek yarıyıllar: {semesters}")
        if debug_mode:
            print(f"BLM ders sayısı: {len(blm_courses)}")
            print(f"YZM ders sayısı: {len(yzm_courses)}")
        
        # Ortak dersleri bul (her iki bölüme de ait olan dersler)
        common_courses = []
        blm_course_codes = {course.code for course in blm_courses}
        yzm_course_codes = {course.code for course in yzm_courses}
        
        for course in all_courses:
            # Birden fazla bölüme ait olan dersleri belirle
            departments = [d.id for d in course.departments]
            if blm_dept.id in departments and yzm_dept.id in departments:
                common_courses.append(course)
        
        if debug_mode:
            print(f"Ortak ders sayısı: {len(common_courses)}")
            
            # Debug: Tüm dersleri yazdır
            print("\n=== TÜM DERSLER ===")
            print("BLM Dersleri:")
            for course in blm_courses:
                # Her dersin bölümlerini göster
                dept_codes = [d.code for d in course.departments]
                print(f"- {course.code} - {course.name} (Yarıyıl: {course.semester}, Bölümler: {', '.join(dept_codes)})")
            
            print("\nYZM Dersleri:")
            for course in yzm_courses:
                # Her dersin bölümlerini göster
                dept_codes = [d.code for d in course.departments]
                print(f"- {course.code} - {course.name} (Yarıyıl: {course.semester}, Bölümler: {', '.join(dept_codes)})")
            
            print("\n=== ORTAK DERSLER ===")
            for course in common_courses:
                # Ortak dersleri göster
                dept_codes = [d.code for d in course.departments]
                print(f"- {course.code} - {course.name} (Yarıyıl: {course.semester}, Bölümler: {', '.join(dept_codes)})")
            
        # Yerleştirilen derslerin izlenmesi için set
        scheduled_courses = set()
        
        # Dersliğe yerleştirmeye yardımcı fonksiyon
        def place_course(course, day, time_slot):
            start_time, end_time = time_slot
            
            # Ders için öğretim üyesi atanmış mı kontrol et
            if course.instructor_id:
                instructor = User.query.get(course.instructor_id)
                
                # Öğretim üyesinin bu gün ve saatte müsait olmama durumu var mı kontrol et
                unavailable_times = UnavailableTime.query.filter(
                    UnavailableTime.instructor_id == course.instructor_id,
                    UnavailableTime.day == day,
                    ((UnavailableTime.start_time <= start_time) & (UnavailableTime.end_time > start_time)) | 
                    ((UnavailableTime.start_time < end_time) & (UnavailableTime.end_time >= end_time)) |
                    ((UnavailableTime.start_time >= start_time) & (UnavailableTime.end_time <= end_time))
                ).all()
                
                if unavailable_times:
                    if debug_mode:
                        print(f'Öğretim üyesi ({instructor.name}) bu zaman diliminde müsait değil!')
                    return False, None
                
                # Öğretim üyesinin bu zaman diliminde başka dersi var mı kontrol et
                instructor_conflicts = Schedule.query.join(Course).filter(
                    Schedule.day == day,
                    Schedule.start_time < end_time,
                    Schedule.end_time > start_time,
                    Course.instructor_id == course.instructor_id
                ).all()
                
                if instructor_conflicts:
                    if debug_mode:
                        conflict_details = []
                        for conflict in instructor_conflicts:
                            conflict_course = Course.query.get(conflict.course_id)
                            conflict_classroom = Classroom.query.get(conflict.classroom_id)
                            conflict_details.append(f"{conflict_course.code} ({conflict_classroom.code}, {conflict.start_time}-{conflict.end_time})")
                        conflict_message = ", ".join(conflict_details)
                        print(f'Öğretim üyesi ({instructor.name}) başka derste meşgul: {conflict_message}')
                    return False, None
            
            # Bu ders için yarıyıldaki diğer derslerin çakışma kontrolü
            # Dersin yarıyılını al
            semester = course.semester
            
            # Bu yarıyıldaki ve aynı bölümlerdeki diğer dersleri bul
            # Şu an çoka-çok ilişki olduğu için her bölüm için ayrı kontrol yapılmalı
            for dept in course.departments:
                # Bu ders programdaki çakışmaları kontrol et (aynı yarıyıl ve bölüm)
                conflict_schedules = db.session.query(Schedule).join(Course).join(
                    course_department, Course.id == course_department.c.course_id
                ).filter(
                    Schedule.day == day,
                    Schedule.start_time == start_time,
                    Schedule.end_time == end_time,
                    Course.semester == semester,
                    course_department.c.department_id == dept.id
                ).all()
                
                if conflict_schedules:
                    if debug_mode:
                        conflict_courses = []
                        for schedule in conflict_schedules:
                            conflict_course = Course.query.get(schedule.course_id)
                            conflict_courses.append(f"{conflict_course.code}")
                        print(f"ÇAKIŞMA: {course.code} dersi {day} günü {start_time}-{end_time} saatinde {dept.code} bölümü {semester}. yarıyıldaki şu derslerle çakışıyor: {', '.join(conflict_courses)}")
                    return False, None
            
            # Uygun derslik bul
            suitable_classroom = None
            
            if course.practice > 0:  # Uygulamalı ders
                lab_classrooms = [c for c in classrooms if c.type == 'LAB']
                free_classrooms = []
                
                for classroom in lab_classrooms:
                    is_occupied = Schedule.query.filter_by(
                        classroom_id=classroom.id,
                        day=day,
                        start_time=start_time,
                        end_time=end_time
                    ).first()
                    
                    if not is_occupied:
                        free_classrooms.append(classroom)
                
                if free_classrooms:
                    suitable_classroom = random.choice(free_classrooms)
            
            if not suitable_classroom:  # Normal derslik
                normal_classrooms = [c for c in classrooms if c.type == 'NORMAL']
                free_classrooms = []
                
                for classroom in normal_classrooms:
                    is_occupied = Schedule.query.filter_by(
                        classroom_id=classroom.id,
                        day=day,
                        start_time=start_time,
                        end_time=end_time
                    ).first()
                    
                    if not is_occupied:
                        free_classrooms.append(classroom)
                
                if free_classrooms:
                    suitable_classroom = random.choice(free_classrooms)
            
            if not suitable_classroom:
                if debug_mode:
                    print(f"DERSLİK BULUNAMADI: {course.code} dersi için {day} günü {start_time}-{end_time} saatinde uygun derslik yok!")
                return False, None
            
            # Programa ekle
            schedule = Schedule(
                course_id=course.id,
                classroom_id=suitable_classroom.id,
                day=day,
                start_time=start_time,
                end_time=end_time
            )
            db.session.add(schedule)
            if debug_mode:
                print(f"YERLEŞTİRİLDİ: {course.code} dersi {day} günü {start_time}-{end_time} saatlerinde {suitable_classroom.code} dersliğine yerleştirildi.")
            return True, suitable_classroom
        
        # 1. ADIM: ORTAK DERSLERİ PROGRAMLA
        print("\n=== ORTAK DERSLER YERLEŞTİRİLİYOR ===")
        for course in common_courses:
            # Bu ders daha önce programlanmış mı kontrol et
            if course.code in scheduled_courses:
                if debug_mode:
                    print(f"ATLANDI: {course.code} dersi zaten programlanmış.")
                continue
                
            placed = False
            attempts = 0
            max_attempts = 100
            
            if debug_mode:
                print(f"ORTAK DERS: {course.code} - {course.name} (Bölümler: {', '.join([d.code for d in course.departments])})")
            
            while not placed and attempts < max_attempts:
                day = random.choice(days)
                time_slot = random.choice(time_slots)
                
                result, _ = place_course(course, day, time_slot)
                if result:
                    scheduled_courses.add(course.code)
                    placed = True
                
                attempts += 1
            
            if not placed:
                print(f"UYARI: {course.code} dersi için uygun zaman dilimi bulunamadı.")
        
        # 2. ADIM: ORTAK OLMAYAN BLM DERSLERİNİ PROGRAMLA
        print("\n=== BLM BÖLÜMÜ DERSLERİ YERLEŞTİRİLİYOR ===")
        for course in blm_courses:
            # Sadece BLM'ye ait olan dersleri programla
            if course in common_courses or course.code in scheduled_courses:
                continue
                
            placed = False
            attempts = 0
            max_attempts = 100
            
            while not placed and attempts < max_attempts:
                day = random.choice(days)
                time_slot = random.choice(time_slots)
                
                result, _ = place_course(course, day, time_slot)
                if result:
                    scheduled_courses.add(course.code)
                    placed = True
                
                attempts += 1
            
            if not placed:
                print(f"UYARI: {course.code} dersi için uygun zaman dilimi bulunamadı.")
        
        # 3. ADIM: ORTAK OLMAYAN YZM DERSLERİNİ PROGRAMLA
        print("\n=== YZM BÖLÜMÜ DERSLERİ YERLEŞTİRİLİYOR ===")
        for course in yzm_courses:
            # Sadece YZM'ye ait olan dersleri programla
            if course in common_courses or course.code in scheduled_courses:
                continue
                
            placed = False
            attempts = 0
            max_attempts = 100
            
            while not placed and attempts < max_attempts:
                day = random.choice(days)
                time_slot = random.choice(time_slots)
                
                result, _ = place_course(course, day, time_slot)
                if result:
                    scheduled_courses.add(course.code)
                    placed = True
                
                attempts += 1
            
            if not placed:
                print(f"UYARI: {course.code} dersi için uygun zaman dilimi bulunamadı.")
        
        db.session.commit()
        
        # Özet bilgiler
        print(f"\n=== PROGRAM OLUŞTURMA TAMAMLANDI ===")
        print(f"Toplam programlanan ders sayısı: {len(scheduled_courses)}")
        if debug_mode:
            print(f"Toplam ortak ders sayısı: {len(common_courses)}")
        
        return True, f"{term_name} dönemi için ders programı başarıyla oluşturuldu."
        
    except Exception as e:
        db.session.rollback()
        print(f"Hata: {str(e)}")
        import traceback
        traceback.print_exc()
        return False, f"Ders programı oluşturulurken bir hata oluştu: {str(e)}"

@app.route('/generate_schedule', methods=['POST'])
@admin_required
def generate_schedule_route():
    """
    Otomatik ders programı oluşturma endpoint'i
    """
    # Seçilen dönemi al
    term = request.form.get('term')
    
    if term:
        success, message = generate_schedule(term)
        if success:
            flash(message, 'success')
        else:
            flash(message, 'error')
    else:
        flash("Lütfen bir dönem seçiniz.", 'error')
        
    return redirect(url_for('view_schedule'))

# Öğretim üyesi kişisel ders programı görüntüleme sayfası
@app.route('/my_schedule')
@login_required
def my_schedule():
    """
    Öğretim üyesinin kendi ders programını görüntüleme sayfası
    Sadece giriş yapmış öğretim üyesi kendi derslerini görüntüleyebilir
    """
    if current_user.role != 'instructor':
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    # Haftanın günleri
    days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    
    # Saat dilimleri (09:00, 10:00, ... 17:00)
    hours = []
    for i in range(9, 18):
        hours.append(f"{i:02d}:00")
    
    # Öğretim üyesinin verdiği dersleri bul
    instructor_courses = Course.query.filter_by(instructor_id=current_user.id).all()
    
    # Ders programı öğelerini bul
    schedule_items = []
    for course in instructor_courses:
        items = Schedule.query.filter_by(course_id=course.id).all()
        schedule_items.extend(items)
    
    # Öğretim üyesinin müsait olmadığı zamanları getir
    unavailable_times = UnavailableTime.query.filter_by(instructor_id=current_user.id).all()
    
    return render_template('my_schedule.html',
                          schedule_items=schedule_items,
                          days=days,
                          hours=hours,
                          unavailable_times=unavailable_times)

# Müsait olmama durumu eklemek için AJAX endpoint
@app.route('/add_unavailable_time', methods=['POST'])
@login_required
def add_unavailable_time():
    """
    Öğretim üyesi için müsait olmama durumu ekler
    JSON verisi içinde day, start_time ve end_time bilgilerini bekler
    """
    if current_user.role != 'instructor':
        return jsonify(success=False, error="Bu işlem için yetkiniz yok.")
    
    try:
        data = request.get_json()
        day = data.get('day')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        # Aynı zaman diliminde başka bir kayıt var mı kontrol et
        existing = UnavailableTime.query.filter_by(
            instructor_id=current_user.id,
            day=day,
            start_time=start_time,
            end_time=end_time
        ).first()
        
        if existing:
            return jsonify(success=True, message="Bu zaman dilimi zaten müsait değil olarak işaretlenmiş.")
        
        # Yeni müsait olmama durumu oluştur
        unavailable_time = UnavailableTime(
            instructor_id=current_user.id,
            day=day,
            start_time=start_time,
            end_time=end_time,
            reason="Tablo üzerinden ayarlandı"
        )
        
        db.session.add(unavailable_time)
        db.session.commit()
        
        return jsonify(success=True)
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e))

# Müsait olmama durumunu kaldırmak için AJAX endpoint
@app.route('/remove_unavailable_time', methods=['POST'])
@login_required
def remove_unavailable_time():
    """
    Öğretim üyesi için müsait olmama durumunu kaldırır
    JSON verisi içinde day, start_time ve end_time bilgilerini bekler
    """
    if current_user.role != 'instructor':
        return jsonify(success=False, error="Bu işlem için yetkiniz yok.")
    
    try:
        data = request.get_json()
        day = data.get('day')
        start_time = data.get('start_time')
        end_time = data.get('end_time')
        
        # İlgili müsait olmama kaydını bul
        unavailable_time = UnavailableTime.query.filter_by(
            instructor_id=current_user.id,
            day=day,
            start_time=start_time,
            end_time=end_time
        ).first()
        
        if unavailable_time:
            db.session.delete(unavailable_time)
            db.session.commit()
            return jsonify(success=True)
        
        return jsonify(success=False, error="Bu zaman dilimi için müsait olmama kaydı bulunamadı.")
    except Exception as e:
        db.session.rollback()
        return jsonify(success=False, error=str(e))

# Öğretim üyelerinin ders programlarını görüntüleme sayfası
@app.route('/instructor_schedules')
@app.route('/instructor_schedules/<int:instructor_id>')
@login_required
def instructor_schedules(instructor_id=None):
    """
    Tüm öğretim üyelerinin ders programlarını görüntüleme sayfası
    Giriş yapmış tüm kullanıcılar bu sayfayı görüntüleyebilir
    """
    if current_user.role != 'instructor' and current_user.role != 'admin':
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    # Tüm öğretim üyelerini getir
    instructors = User.query.filter_by(role='instructor').order_by(User.name).all()
    
    # Haftanın günleri
    days = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
    
    # Saat dilimleri (09:00, 10:00, ... 17:00)
    hours = []
    for i in range(9, 18):
        hours.append(f"{i:02d}:00")
    
    schedule_items = []
    selected_instructor = None
    
    # Eğer bir öğretim üyesi seçilmişse
    if instructor_id:
        selected_instructor = User.query.filter_by(id=instructor_id, role='instructor').first_or_404()
        
        # Seçilen öğretim üyesinin verdiği dersleri bul
        instructor_courses = Course.query.filter_by(instructor_id=instructor_id).all()
        
        # Ders programı öğelerini bul
        for course in instructor_courses:
            items = Schedule.query.filter_by(course_id=course.id).all()
            schedule_items.extend(items)
    else:
        # Öğretim üyesi seçilmemişse ve kullanıcı bir öğretim üyesi ise, kendi programını göster
        if current_user.role == 'instructor':
            selected_instructor = current_user
            instructor_courses = Course.query.filter_by(instructor_id=current_user.id).all()
            
            for course in instructor_courses:
                items = Schedule.query.filter_by(course_id=course.id).all()
                schedule_items.extend(items)
    
    return render_template('instructor_schedules.html',
                          instructors=instructors,
                          selected_instructor=selected_instructor,
                          schedule_items=schedule_items,
                          days=days,
                          hours=hours)


# Kişisel ders programını Excel'e aktarma
@app.route('/export_my_schedule')
@login_required
def export_my_schedule():
    """
    Öğretim üyesinin kendi ders programını Excel'e aktarma
    """
    if current_user.role != 'instructor':
        flash('Bu sayfaya erişim yetkiniz yok.', 'error')
        return redirect(url_for('index'))
    
    # Öğretim üyesinin verdiği dersleri bul
    instructor_courses = Course.query.filter_by(instructor_id=current_user.id).all()
    
    # Ders programı öğelerini bul
    schedule_items = []
    for course in instructor_courses:
        items = Schedule.query.filter_by(course_id=course.id).all()
        schedule_items.extend(items)
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Ders Programı"
    
    # Başlık satırı
    headers = ["Gün", "Ders Kodu", "Ders Adı", "Saat", "Derslik", "Bölüm"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Verileri ekle
    row_num = 2
    for item in sorted(schedule_items, key=lambda x: (x.day, x.start_time)):
        course = Course.query.get(item.course_id)
        classroom = Classroom.query.get(item.classroom_id)
        
        # Bölüm adlarını birleştir
        departments = ", ".join([dept.code for dept in course.departments])
        
        ws.cell(row=row_num, column=1).value = item.day
        ws.cell(row=row_num, column=2).value = course.code
        ws.cell(row=row_num, column=3).value = course.name
        ws.cell(row=row_num, column=4).value = f"{item.start_time}-{item.end_time}"
        ws.cell(row=row_num, column=5).value = classroom.code
        ws.cell(row=row_num, column=6).value = departments
        
        row_num += 1
    
    # Sütun genişliklerini ayarla
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    
    # Dosyayı geçici olarak kaydet ve kullanıcıya gönder
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_file = tmp.name
    
    return send_file(tmp_file, 
                    as_attachment=True, 
                    download_name=f"{current_user.name}_Ders_Programi.xlsx")

# Uygulama başlangıç kontrollerini yap ve sunucuyu başlat
if __name__ == '__main__':
    """
    Uygulama başlatıldığında çalışır
    - Veritabanı tabloları oluşturulur (yoksa)
    - Admin kullanıcısı oluşturulur (yoksa)
    - Flask geliştirme sunucusu başlatılır
    """
    with app.app_context():
        # Veritabanı tablolarını oluştur
        db.create_all()
        
        # Eksik sütunları ekle (migrasyon)
        try:
            # Course tablosunda instructor_id sütunu var mı kontrol et
            inspector = inspect(db.engine)
            
            # Course tablosuna instructor_id ekle
            if 'instructor_id' not in [c['name'] for c in inspector.get_columns('courses')]:
                with db.engine.begin() as conn:
                    conn.execute(text("ALTER TABLE courses ADD COLUMN instructor_id INTEGER REFERENCES users(id)"))
                print("courses tablosuna instructor_id sütunu eklendi.")
            
            # Diğer eksik sütunları da kontrol et ve ekle
            if 'semester' not in [c['name'] for c in inspector.get_columns('courses')]:
                with db.engine.begin() as conn:
                    conn.execute(text("ALTER TABLE courses ADD COLUMN semester INTEGER DEFAULT 1"))
                print("courses tablosuna semester sütunu eklendi.")
        except Exception as e:
            print(f"Migrasyon hatası: {str(e)}")
        
        # Admin kullanıcısı oluştur (yoksa)
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(username='admin', password='admin123', role='admin', name='Sistem Yöneticisi')
            db.session.add(admin)
            db.session.commit()
            print("Admin kullanıcısı oluşturuldu. Kullanıcı adı: admin, Şifre: admin123")
    
    # Geliştirme sunucusunu başlat
    app.run(debug=True) 